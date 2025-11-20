import openpyxl
from openpyxl.styles import Color
import pandas as pd
import numpy as np
import glob
import os

def extract_highlighted_coordinates(file_path, sheet_name):
    """
    从Excel文件中提取黄色高亮区域的坐标
    
    参数:
        file_path: Excel文件路径
        sheet_name: sheet名称
    
    返回:
        highlighted_coords: 高亮单元格的坐标列表 [(行索引, 列索引), ...]
        matrix_data: 矩阵数据（不含行列序号）
    """
    # 加载工作簿
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    
    # 获取矩阵数据和高亮坐标
    highlighted_coords = []
    matrix_data = []
    
    # 读取数据（跳过第一行和第一列的序号）
    max_row = ws.max_row
    max_col = ws.max_column
    
    # 提取行序号（第一列，从第二行开始）
    row_indices = []
    for i in range(2, max_row + 1):
        cell = ws.cell(row=i, column=1)
        row_indices.append(cell.value)
    
    # 提取列序号（第一行，从第二列开始）
    col_indices = []
    for j in range(2, max_col + 1):
        cell = ws.cell(row=1, column=j)
        col_indices.append(cell.value)
    
    # 遍历矩阵区域（从第2行第2列开始）
    for i in range(2, max_row + 1):
        row_data = []
        for j in range(2, max_col + 1):
            cell = ws.cell(row=i, column=j)
            row_data.append(cell.value)
            
            # 检查单元格是否有黄色高亮
            if cell.fill.start_color:
                color = cell.fill.start_color.rgb
                # 检测黄色（通常是FFFFFF00或类似的颜色）
                if color and isinstance(color, str):
                    # 移除可能的透明度前缀
                    if len(color) == 8:
                        color = color[2:]
                    
                    # 判断是否为黄色系（可以根据实际情况调整阈值）
                    if is_yellow(color):
                        # 使用矩阵坐标（基于行列序号）
                        row_idx = row_indices[i - 2]  # i-2 因为从第2行开始
                        col_idx = col_indices[j - 2]  # j-2 因为从第2列开始
                        highlighted_coords.append((row_idx, col_idx))
        
        matrix_data.append(row_data)
    
    wb.close()
    
    return highlighted_coords, matrix_data, row_indices, col_indices


def is_yellow(rgb_hex):
    """
    判断RGB颜色是否为黄色系
    """
    try:
        # 转换为RGB值
        r = int(rgb_hex[0:2], 16)
        g = int(rgb_hex[2:4], 16)
        b = int(rgb_hex[4:6], 16)
        
        # 黄色判断：R和G值较高，B值较低
        # 可以根据实际情况调整阈值
        return r > 200 and g > 200 and b < 100
    except:
        return False


def print_results(highlighted_coords, matrix_data, row_indices, col_indices):
    """
    打印结果
    """
    print("=" * 50)
    print("高亮区域的坐标：")
    print("=" * 50)
    for coord in highlighted_coords:
        print(f"行索引: {coord[0]}, 列索引: {coord[1]}")
    
    print("\n" + "=" * 50)
    print(f"共找到 {len(highlighted_coords)} 个高亮单元格")
    print("=" * 50)
    
    # 可选：将矩阵转换为DataFrame便于查看
    df = pd.DataFrame(matrix_data, index=row_indices, columns=col_indices)
    print("\n矩阵数据预览：")
    print(df)


def read_txt_matrix(txt_file_path, delimiter=None):
    """
    从txt文件中读取矩阵
    
    参数:
        txt_file_path: txt文件路径
        delimiter: 分隔符（None表示自动检测空白字符）
    
    返回:
        matrix: numpy数组形式的矩阵
    """
    try:
        # 读取txt文件
        matrix = np.loadtxt(txt_file_path, delimiter=delimiter)
        print(f"成功读取txt矩阵，维度: {matrix.shape}")
        return matrix
    except Exception as e:
        print(f"读取txt文件出错: {str(e)}")
        # 尝试用pandas读取
        try:
            df = pd.read_csv(txt_file_path, sep=delimiter, header=None, engine='python')
            matrix = df.values
            print(f"成功读取txt矩阵（使用pandas），维度: {matrix.shape}")
            return matrix
        except Exception as e2:
            print(f"使用pandas也失败: {str(e2)}")
            return None


def extract_submatrix(matrix, coordinates):
    """
    根据坐标从矩阵中提取子矩阵
    
    参数:
        matrix: 完整矩阵（numpy数组）
        coordinates: 坐标列表 [(行索引, 列索引), ...]
                    注意：索引从1开始（Excel风格）
    
    返回:
        submatrix: 提取的子矩阵元素列表
        coord_list: 对应的坐标列表
    """
    submatrix_elements = []
    valid_coords = []
    
    for row_idx, col_idx in coordinates:
        try:
            # Excel索引从1开始，转换为Python索引（从0开始）
            # 假设Excel中的行列序号对应txt矩阵的实际位置
            r = int(row_idx) - 1  # 转换为0-based索引
            c = int(col_idx) - 1
            
            # 检查索引是否在范围内
            if 0 <= r < matrix.shape[0] and 0 <= c < matrix.shape[1]:
                element = matrix[r, c]
                submatrix_elements.append(element)
                valid_coords.append((row_idx, col_idx))
                print(f"提取位置 ({row_idx}, {col_idx}): {element}")
            else:
                print(f"警告: 坐标 ({row_idx}, {col_idx}) 超出矩阵范围")
        except Exception as e:
            print(f"提取坐标 ({row_idx}, {col_idx}) 时出错: {str(e)}")
    
    return submatrix_elements, valid_coords


def save_submatrix_to_txt(elements, coordinates, output_file, format_type='list'):
    """
    将子矩阵保存到txt文件
    
    参数:
        elements: 提取的元素列表
        coordinates: 对应的坐标列表
        output_file: 输出文件路径
        format_type: 'list'(列表格式) 或 'matrix'(紧凑矩阵格式)
    """
    with open(output_file, 'w', encoding='utf-8') as f:
        if format_type == 'list':
            # 格式1：每行包含坐标和值
            f.write("# 行索引\t列索引\t值\n")
            for (r, c), val in zip(coordinates, elements):
                f.write(f"{r}\t{c}\t{val}\n")
        
        elif format_type == 'values_only':
            # 格式2：只保存值（每行一个）
            for val in elements:
                f.write(f"{val}\n")
        
        elif format_type == 'matrix':
            # 格式3：紧凑矩阵格式（只包含实际存在的行和列）
            # 找出所有唯一的行和列索引
            rows = sorted(set(r for r, c in coordinates))
            cols = sorted(set(c for r, c in coordinates))
            
            # 创建字典便于查找
            coord_dict = {(r, c): val for (r, c), val in zip(coordinates, elements)}
            
            # 写入矩阵（只输出存在数据的行和列）
            for r in rows:
                row_values = []
                for c in cols:
                    if (r, c) in coord_dict:
                        row_values.append(str(coord_dict[(r, c)]))
                
                # 只有当该行有数据时才写入
                if row_values:
                    f.write('\t'.join(row_values) + '\n')
        
        elif format_type == 'matrix_dense':
            # 格式4：完全紧凑的矩阵格式（按行组织，每行只包含该行实际存在的元素）
            # 按行分组
            from collections import defaultdict
            row_data = defaultdict(list)
            
            for (r, c), val in zip(coordinates, elements):
                row_data[r].append((c, val))
            
            # 按行索引排序
            for r in sorted(row_data.keys()):
                # 按列索引排序该行的数据
                sorted_cols = sorted(row_data[r], key=lambda x: x[0])
                # 只输出值
                row_values = [str(val) for _, val in sorted_cols]
                f.write('\t'.join(row_values) + '\n')
        
        elif format_type == 'matrix_with_indices':
            # 格式5：带行列索引的紧凑矩阵（首行为列索引，首列为行索引）
            rows = sorted(set(r for r, c in coordinates))
            cols = sorted(set(c for r, c in coordinates))
            
            # 创建字典便于查找
            coord_dict = {(r, c): val for (r, c), val in zip(coordinates, elements)}
            
            # 写入列索引（首行）
            f.write('\t' + '\t'.join(map(str, cols)) + '\n')
            
            # 写入每行数据（首列为行索引）
            for r in rows:
                row_values = [str(r)]  # 首列为行索引
                for c in cols:
                    if (r, c) in coord_dict:
                        row_values.append(str(coord_dict[(r, c)]))
                    else:
                        row_values.append('')  # 空白表示无数据
                
                # 只有当该行有实际数据时才写入
                if any(row_values[1:]):  # 检查除了行索引外是否有数据
                    f.write('\t'.join(row_values) + '\n')
    
    print(f"子矩阵已保存到: {output_file}")


def main_test(excel_file, sheet_name, txt_file, output_file, txt_delimiter=None):
    """
    主函数：完整流程
    
    参数:
        excel_file: Excel文件路径
        sheet_name: Excel的sheet名称
        txt_file: txt矩阵文件路径
        output_file: 输出文件路径
        txt_delimiter: txt文件的分隔符
    """
    print("=" * 60)
    print("步骤1: 从Excel提取高亮坐标")
    print("=" * 60)
    
    # 提取高亮坐标
    coordinates,_,_,_ = extract_highlighted_coordinates(excel_file, sheet_name)
    print(f"找到 {len(coordinates)} 个高亮单元格")
    print(f"坐标列表: {coordinates[:10]}..." if len(coordinates) > 10 else f"坐标列表: {coordinates}")
    
    print("\n" + "=" * 60)
    print("步骤2: 读取txt矩阵")
    print("=" * 60)
    
    # 读取txt矩阵
    matrix = read_txt_matrix(txt_file, delimiter=txt_delimiter)
    if matrix is None:
        print("无法读取txt矩阵,程序终止")
        return
    
    print("\n" + "=" * 60)
    print("步骤3: 提取子矩阵元素")
    print("=" * 60)
    
    # 提取子矩阵
    elements, valid_coords = extract_submatrix(matrix, coordinates)
    print(f"成功提取 {len(elements)} 个元素")
    
    print("\n" + "=" * 60)
    print("步骤4: 保存结果（多种格式）")
    print("=" * 60)
    
    # # 格式1：带坐标的列表
    # save_submatrix_to_txt(elements, valid_coords, 
    #                      output_file.replace('.txt', '_with_coords.txt'), 
    #                      format_type='list')
    
    # # 格式2：只有值
    # save_submatrix_to_txt(elements, valid_coords, 
    #                      output_file.replace('.txt', '_values_only.txt'), 
    #                      format_type='values_only')
    
    # # 格式3：紧凑矩阵格式（不补0，保持原有行列结构）
    # save_submatrix_to_txt(elements, valid_coords, 
    #                      output_file.replace('.txt', '_matrix_compact.txt'), 
    #                      format_type='matrix')
    
    # 格式4：完全紧凑格式（每行只包含该行的数据）
    save_submatrix_to_txt(elements, valid_coords, 
                         output_file.replace('.txt', '_matrix_dense.txt'), 
                         format_type='matrix_dense')
    
    # # 格式5：带行列索引的紧凑矩阵
    # save_submatrix_to_txt(elements, valid_coords, 
    #                      output_file.replace('.txt', '_matrix_indexed.txt'), 
    #                      format_type='matrix_with_indices')
    
    print("\n" + "=" * 60)
    print("处理完成！生成了以下文件：")
    # print("  1. *_with_coords.txt - 坐标+值列表格式")
    # print("  2. *_values_only.txt - 仅值列表")
    # print("  3. *_matrix_compact.txt - 紧凑矩阵（保持原行列结构）")
    print("  4. *_matrix_dense.txt - 完全紧凑矩阵（按行紧密排列）")
    # print("  5. *_matrix_indexed.txt - 带索引的紧凑矩阵")
    print("=" * 60)

def extract_matrix_from_folder(matrix_dir, output_dir, coordinates):
    # 创建输出目录（如果不存在）
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"创建输出目录: {output_dir}")

    # 获取所有.txt文件
    txt_files = glob.glob(os.path.join(matrix_dir, "*.txt"))
    txt_files.sort()  # 排序以便按顺序处理

    print(f"在 {matrix_dir} 中找到 {len(txt_files)} 个txt文件")

    if len(txt_files) == 0:
        print("警告：未找到任何txt文件！")
    else:
        # 批量处理每个txt文件
        success_count = 0
        fail_count = 0
        
        for idx, txt_file in enumerate(txt_files, 1):
            # 获取文件名（不含路径）
            file_name = os.path.basename(txt_file)
            print(f"\n[{idx}/{len(txt_files)}] 处理文件: {file_name}")
            
            # 生成输出文件名
            base_name = os.path.splitext(file_name)[0]  # 去除.txt后缀
            output_file = os.path.join(output_dir, f"{base_name}_submatrix.txt")
            
            try:
                # Step 2. read txt file, extract matrix
                print(f"  - 读取矩阵...")
                matrix = read_txt_matrix(txt_file, delimiter=txt_delimiter)
                if matrix is None:
                    print(f"  ✗ 无法读取txt矩阵: {file_name}")
                    fail_count += 1
                    continue
                
                # Step 3. extract submatrix
                print(f"  - 提取子矩阵...")
                elements, valid_coords = extract_submatrix(matrix, coordinates)
                print(f"  - 成功提取 {len(elements)} 个元素")
                
                # Step 4. save submatrix
                print(f"  - 保存结果...")
                save_submatrix_to_txt(elements, valid_coords, 
                                    output_file, format_type='matrix_dense')
                
                print(f"  ✓ 处理成功: {base_name}_submatrix.txt")
                success_count += 1
                
            except Exception as e:
                print(f"  ✗ 处理失败: {file_name}")
                print(f"    错误信息: {str(e)}")
                fail_count += 1
                import traceback
                traceback.print_exc()

        # 打印汇总信息
        print("\n" + "=" * 60)
        print("批量处理完成！")
        print("=" * 60)
        print(f"总文件数: {len(txt_files)}")
        print(f"成功处理: {success_count}")
        print(f"失败数量: {fail_count}")
        print(f"输出目录: {output_dir}")
        print("=" * 60)


# 使用示例
def main_process_submatrix_extraction():
    # 配置参数
    data_dir = "/mnt/c/Works/ws/shoufa2025/data/matrix"
    excel_file = os.path.join(data_dir, "Excel_Original_Data_rest_no_editing_pre_surgery.xlsx")  # 子图高亮Excel文件路径
    sheet_name = "sub_0001"    # Sheet名称

    # txt分隔符：None(空白字符), '\t'(制表符), ','(逗号), ' '(空格)等
    txt_delimiter = None

    # Step 1. highlighted coordinates
    print("=" * 60)
    print("Step 1: 提取高亮坐标")
    print("=" * 60)
    coordinates, _, _, _ = extract_highlighted_coordinates(excel_file, sheet_name)
    print(f"找到 {len(coordinates)} 个高亮单元格")
    print(f"坐标列表: {coordinates[:10]}..." if len(coordinates) > 10 else f"坐标列表: {coordinates}")

    # Step 1.5. 从matrix_dir中读取.txt文件路径列表，并批量处理
    print("\n" + "=" * 60)
    print("Step 1.5: 批量处理所有txt文件")
    print("=" * 60)

    matrix_dir = os.path.join(data_dir, "post_surgery_original_GretnaSFCMatrixZ")
    output_dir = os.path.join(data_dir, "processed", "post")

    extract_matrix_from_folder(matrix_dir, output_dir, coordinates)


if __name__ == "__main__":
    # run submatrix extraction, from folders
    main_process_submatrix_extraction()

    # run category arrangement, by 首发入组MRI对照表, sheet=剔除无MRI, key=CAM评分术后, subname=MRI排序
    